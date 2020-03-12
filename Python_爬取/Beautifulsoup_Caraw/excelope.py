import openpyxl as op
#创建一个工作簿和工作表
# wb = op.Workbook() #创建一个工作薄
# wb.create_sheet('Top250')#创建一个工作表
# wb.save('豆瓣.xlsx')#保存
# wb.close()#关闭

#打开工作簿进行读
# 打开工作簿：wb = openpyxl.load_workbook('cases.xlsx')
# 选取表单：sh = wb['Sheet1'
# 读取第一行、第一列的数据：ce = sh.cell(row = 1,column = 1)
# 按行读取数据：row_data = list(sh.rows)
# 关闭工作薄：wb.close()
# 按列读取数据：columns_data = list(sh.columns)

#写入工作表（关闭状态）

class WriteXlsx:

    def write(self,dataList=[]):
        wb = op.load_workbook('豆瓣.xlsx')
        sh = wb["Top250"]
        title = "豆瓣Top250电影"
        sh["A1"] = title
        print(len(dataList))
        for data,i in zip(dataList,range(len(dataList))):
            sh.cell(row=i+2,column=1,value=data['作品'])
            sh.cell(row=i+2, column=2, value=data['链接'])
            sh.cell(row=i+2, column=3, value=data['导演'])
            sh.cell(row=i+2, column=4, value=data['时间和类型'])
            sh.cell(row=i+2, column=5, value=data['主要内容'])
        wb.save('豆瓣.xlsx')
        wb.close()
